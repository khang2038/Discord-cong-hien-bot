import discord
from discord.ext import commands
from discord import app_commands
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from keep_alive import keep_alive
from dotenv import load_dotenv

# ========== CẤU HÌNH ==========
load_dotenv()
TOKEN = os.getenv("DISCORD_TOKEN")  # ← Lấy token từ .env
FILE_NAME = "conghien.xlsx"
# ==============================

intents = discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix="!", intents=intents)
tree = bot.tree

# Tạo file nếu chưa có
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contributions"
    ws.append(["User ID", "Username", "Time", "Content"])
    wb.save(FILE_NAME)

# Ghi nhận cống hiến
@tree.command(name="conghien", description="Ghi nhận một đóng góp mới")
@app_commands.describe(noi_dung="Nội dung bạn muốn đóng góp")
async def conghien(interaction: discord.Interaction, noi_dung: str):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([
        interaction.user.id,
        str(interaction.user),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        noi_dung
    ])
    wb.save(FILE_NAME)
    await interaction.response.send_message(
        f"✅ Đã ghi nhận đóng góp của {interaction.user.mention}:\n**{noi_dung}**"
    )

# Xuất toàn bộ dữ liệu
@tree.command(name="export", description="Xuất toàn bộ đóng góp")
async def export_all(interaction: discord.Interaction):
    if not os.path.exists(FILE_NAME):
        await interaction.response.send_message("❌ Chưa có dữ liệu.", ephemeral=True)
        return
    await interaction.response.send_message("📂 Đây là toàn bộ đóng góp:", file=discord.File(FILE_NAME), ephemeral=True)

@tree.command(name="export_user", description="Xuất đóng góp của một user cụ thể (bằng cách @mention)")
@app_commands.describe(user="Người dùng bạn muốn xuất báo cáo")
async def export_user(interaction: discord.Interaction, user: discord.User):
    if not os.path.exists(FILE_NAME):
        await interaction.response.send_message("❌ Chưa có dữ liệu.", ephemeral=True)
        return

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "User Contributions"
    new_ws.append(["User ID", "Username", "Time", "Content"])

    target_username = str(user)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[1]) == target_username:
            new_ws.append(row)

    if new_ws.max_row == 1:
        await interaction.response.send_message(f"❌ {user.mention} chưa có đóng góp nào.", ephemeral=True)
        return

    export_filename = f"user_{user.id}_export.xlsx"
    new_wb.save(export_filename)
    await interaction.response.send_message(f"📂 Đây là đóng góp của {user.mention}:", file=discord.File(export_filename), ephemeral=True)

# Khi bot online
@bot.event
async def on_ready():
    await tree.sync()
    print(f"✅ Bot đã sẵn sàng dưới tên {bot.user}!")

keep_alive()

bot.run(TOKEN)
